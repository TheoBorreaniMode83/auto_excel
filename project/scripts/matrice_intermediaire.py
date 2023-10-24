
from openpyxl import load_workbook


class MatriceIntermaidiaire:
    # c'est un tableau de tuple contenant comme informations une dénomination et un type
    # c'est un tableau de tabeau contenat des strings représentant le contenu du fichier excel

    def __init__(self, nomFichier, nomFeuille, plage,header):
        self.header = header
        self.content = []
        self.__extractionTab(nomFichier,nomFeuille,plage)
        self.__gestionErreurConstructeur()
    
    def __gestionErreurConstructeur(self):
        if(len(self.header.header)!=len(self.content[0])):
            raise Exception("MatriceIntermaidiaire fausse")


    def __extractionTab(self, nomFichier, nomFeuille, plage):
        classeur = load_workbook(filename = nomFichier)
        feuille = classeur[nomFeuille]
        for i in range(plage.minRow,plage.maxRow+1):
            self.content.append([])   
            for ii in range(plage.minColumn,plage.maxColumn+1):
                cellule = feuille.cell(row=i, column=ii)
                self.content[-1].append(str(cellule.value))

    def __subFunctionAfficheMatrice(self,chaine,tailleVoulu):
        tmp=len(chaine)
        for i in range(len(chaine),tailleVoulu):
            chaine=chaine+' '
        return chaine

    def afficheMatrice(self,separator):

        sizeMemory=[]

        # step 1 : remplir sizeMemory de 0 sa taille sera égal aux nombres de collonnes de la matrice
        for i in range(len(self.header.header)):
            sizeMemory.append(0)
        # step 2 : remplir le tableau size memory contenant la taille de la plus grande chaine de chaque colonnes
        for i in range(len(self.content)):
            for ii in range(len(self.content[0])):
                if len(str(self.header.header[ii]))>sizeMemory[ii]:
                    sizeMemory[ii]=len(self.header.header[ii][0])+len(self.header.header[ii][1])+8
                if len(self.content[i][ii])>sizeMemory[ii]:
                    sizeMemory[ii]= len(self.content[i][ii])
        # step 3.1 : affichage de l'entete du tableau
        for i in range(len(self.header.header)):
            print(self.__subFunctionAfficheMatrice(str(self.header.header[i]),sizeMemory[i]),end="")
            if i!= len(self.header.header):
                print(separator,end="")
        print()
        # step 3.2 : affichage du corps du tableau
        for i in range(len(self.content)):
            for ii in range(len(self.content[0])):
                print(self.__subFunctionAfficheMatrice(self.content[i][ii],sizeMemory[ii]),end="")
                if ii != len(self.content[0]):
                    print(separator,end="")
            print()
        
    def functionOnColumn(self, numberRow, function):
        for i in range(len(self.content)):
            self.content[i][numberRow] = function(self.content[i][numberRow]) 
    
    def echangeRow(self,row1, row2):
        tmp = self.header.header[row1]
        self.header.header[row1] = self.header.header[row2]
        self.header.header[row2] = tmp
        for i in range(len(self.content)):
            tmp = self.content[i][row1]
            self.content[i][row1] = self.content[i][row2]
            self.content[i][row2] = tmp

    def insertRow(self,subTab,index):
        if(len(self.content[0])!=len(subTab)):
            print("error")
            quit()
        self.content.insert(index,subTab)
    
    def deleteRow(self,index):
        self.header.header.pop(index)
        for i in range(len(self.content)):
            self.content[i].pop(index)

    """
    def makeCsv(tableau):
        with open('postTraitement.csv', 'w', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=',',quoting=csv.QUOTE_ALL)# quotechar='|', quoting=csv.QUOTE_MINIMAL
            for i in range(len(tableau)):
                spamwriter.writerow(tableau[i])

    def writeExcel(tab, file, nomFeuille):
        classeur = load_workbook(filename = file)
        feuille = classeur[nomFeuille]
        for i in range(1,len(tab)+1):
            for ii in range(1,len(tab[0])+1):
                cellule = feuille.cell(row=i, column=ii)
                cellule.value = tab[i-1][ii-1]
        classeur.save(file)
    """